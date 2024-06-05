'''import openpyxl

def get_data(path,sheet_name):
    final_list=[]
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_columns+1):
            row_list.append(sheet.cell(r,c).value)
        final_list.append(row_list)
    return final_list'''

'''import openpyxl

def get_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is a header
        data.append(tuple(cell if cell is not None else '' for cell in row))  # Replace None with empty string
    return data'''

import openpyxl

def get_data(path, sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_columns = sheet.max_column

    for r in range(2, total_rows + 1):  
        row_list = []
        for c in range(1, total_columns + 1):
            cell_value = sheet.cell(r, c).value
            row_list.append(cell_value if cell_value is not None else '')  
        final_list.append(tuple(row_list))  
    return final_list

